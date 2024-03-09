from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side


def create_excel_table(box_sizes, box_colors, text_values, font_sizes, file_name):
    # Create a new Workbook
    wb = Workbook()
    ws = wb.active

    # Set column widths based on box sizes
    for col, size in enumerate(box_sizes, start=1):
        ws.column_dimensions[chr(64 + col)].width = size

    # Populate the table with colors and text
    for row, (color_row, text_row, size_row) in enumerate(zip(box_colors, text_values, font_sizes), start=1):
        for col, (color, text, size) in enumerate(zip(color_row, text_row, size_row), start=1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=f"FF{color}", end_color=f"FF{color}", fill_type="solid")
            cell.value = text
            cell.font = Font(size=size, color="000000" if color.lstrip('#') in ['FFFFFF',
                                                                                '000000'] else "000000")  # Set font size and color based on background color

            # Adding border
            border = Border(left=Side(border_style="thin", color='000000'),
                            right=Side(border_style="thin", color='000000'),
                            top=Side(border_style="thin", color='000000'),
                            bottom=Side(border_style="thin", color='000000'))
            cell.border = border

    # Save the workbook
    wb.save(file_name)


# Example usage
box_sizes = [50, 5, 30, 30]  # Adjust sizes as needed
box_colors = [
    ["859EA0", "FFFFFF", "859EA0", "859EA0"],
    ["FFFFFF", "FFFFFF", "FFFFFF", "FFFFFF"],  # Row 1 colors
    ["9DA9AA", "FFFFFF", "FFFFFF", "FFFFFF"],
    ["9DA9AA", "FFFFFF", "FFFFFF", "FFFFFF"],
    ["FFFFFF", "FFFFFF", "FFFFFF", "FFFFFF"],
    ["9DA9AA", "FFFFFF", "FFFFFF", "FFFFFF"],
    ["FFFFFF", "FFFFFF", "FFFFFF", "FFFFFF"],
    ["9DA9AA", "FFFFFF", "FFFFFF", "FFFFFF"],
    ["9DA9AA", "FFFFFF", "FFFFFF", "FFFFFF"],
    ["FFFFFF", "FFFFFF", "FFFFFF", "FFFFFF"],
    ["9DA9AA", "FFFFFF", "FFFFFF", "FFFFFF"],
    ["9DA9AA", "FFFFFF", "FFFFFF", "FFFFFF"],
    ["FFFFFF", "FFFFFF", "FFFFFF", "FFFFFF"],
    ["9DA9AA", "FFFFFF", "FFFFFF", "FFFFFF"],
    ["FFFFFF", "FFFFFF", "FFFFFF", "FFFFFF"],
    ["9DA9AA", "FFFFFF", "FFFFFF", "FFFFFF"],  # Row 3 colors
]
text_values = [
    ["Travel", "", "Ընդհանուր", "xxxx"],  # Row 1 text
    ["", "", "", ""],
    ["Հավագրած Ապ․ Վճար", "", "aaaa", "20000"],
    ["Պայմանագրերի քանակ", "", "210", "100"],
    ["", "", "", ""],
    ["Վաստակած ապ վճար", "", "աաաաա", "աաաաա"],
    ["", "", "", ""],
    ["Հատուցում", "", "աաաաա", "աաաաա"],
    ["Պայմանագրերի քանակ", "", "աաաաա", "աաաաա"],
    ["", "", "", ""],
    ["ՆՉՊՊ", "", "աաաաա", "աաաաա"],
    ["ՆՉՊՊ քանակ", "", "աաաաա", "աաաաա"],
    ["", "", "", ""],
    ["Համախառն", "", "աաաաա", "աաաաա"],
    ["", "", "", ""],
    ["Միջնորդավճար", "", "աաաաա", "աաաաա"],
    # Row 3 text
]
font_sizes = [
    [18, 10, 18, 18],  # Row 1 font sizes
    [10, 10, 12, 12],  # Row 2 font sizes
    [16, 10, 12, 12],
    [16, 10, 12, 12],
    [16, 10, 12, 12],
    [16, 10, 12, 12],
    [16, 10, 12, 12],
    [16, 10, 12, 12],
    [16, 10, 12, 12],
    [16, 10, 12, 12],
    [16, 10, 12, 12],
    [16, 10, 12, 12],
    [16, 10, 12, 12],
    [16, 10, 12, 12],
    [16, 10, 12, 12],
    [16, 10, 12, 12],
]
file_name = "excel_design.xlsx"

create_excel_table(box_sizes, box_colors, text_values, font_sizes, file_name)
